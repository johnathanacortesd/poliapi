import streamlit as st
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, NamedStyle
from openpyxl.utils import get_column_letter
from collections import defaultdict
from difflib import SequenceMatcher
from copy import deepcopy
import datetime
import io
import openai
import re
import time
from unidecode import unidecode
import numpy as np
from sklearn.cluster import AgglomerativeClustering

# --- Configuración de la página y Modelos ---
st.set_page_config(page_title="Análisis de Noticias para Policía Nacional", layout="wide")
OPENAI_MODEL_EMBEDDING = 'text-embedding-3-small'
# MODELO RESTAURADO SEGÚN TU ESPECIFICACIÓN
OPENAI_MODEL_CLASIFICACION = "gpt-4.1-nano-2025-04-14"
MARCA_FIJA = "Policía Nacional de Colombia"
TEXT_TRUNCATION_LIMIT = 2200 # Límite de caracteres a enviar a la API

# ==============================================================================
# SECCIÓN DE AUTENTICACIÓN
# ==============================================================================
def check_password():
    if st.session_state.get("password_correct", False): return True
    st.header("🔐 Acceso Protegido")
    with st.form("password_form"):
        password = st.text_input("Ingresa la contraseña para continuar:", type="password")
        submitted = st.form_submit_button("Ingresar")
        if submitted:
            # Asegúrate de tener "APP_PASSWORD" en tus secrets de Streamlit
            if password == st.secrets.get("APP_PASSWORD", "INVALID_DEFAULT"):
                st.session_state["password_correct"] = True; st.rerun()
            else: st.error("La contraseña es incorrecta.")
    return False

# ==============================================================================
# CLASES Y FUNCIONES DE ANÁLISIS IA
# ==============================================================================

def _limpiar_texto(t):
    return re.sub(r'\s+', ' ', unidecode(str(t)).strip()) if pd.notna(t) else ""

def _get_embedding_cached(texto, cache):
    limpio = _limpiar_texto(texto)
    if not limpio: return None
    if limpio in cache: return cache[limpio]
    try:
        time.sleep(0.02)
        response = openai.Embedding.create(input=[limpio[:TEXT_TRUNCATION_LIMIT]], model=OPENAI_MODEL_EMBEDDING)
        embedding = response['data'][0]['embedding']
        cache[limpio] = embedding
        return embedding
    except Exception as e:
        st.warning(f"API Embedding Warn: {e}")
        return None

class ClasificadorTonoPolicia:
    def __init__(self):
        self.cache_tonos = {}
        self.cache_embeddings = {}
        self.high_similarity_threshold = 0.95

    def _clasificar_gpt(self, texto_resumen):
        limpio = _limpiar_texto(texto_resumen)
        if not limpio: return "Neutro"
        if limpio in self.cache_tonos: return self.cache_tonos[limpio]
        
        prompt = f"""
        Eres un analista de medios experto en la reputación de la "{MARCA_FIJA}".
        Clasifica el TONO de la siguiente noticia en relación con la institución, siguiendo estas reglas estrictas:
        - POSITIVO: La noticia informa sobre gestiones proactivas, actividades beneficiosas, campañas exitosas, jornadas de acompañamiento, consejos de seguridad liderados por la policía, o cualquier acción que genere una percepción claramente favorable.
        - NEGATIVO: La noticia reporta críticas directas a la institución, casos de corrupción, irregularidades, abusos, o ataques contra miembros de la policía.
        - NEUTRO: La noticia es una mención referencial o factual, donde la policía es parte del contexto pero no el foco de un juicio de valor positivo o negativo.

        NOTICIA: --- {limpio[:TEXT_TRUNCATION_LIMIT]} ---

        Responde ÚNICAMENTE con una de las siguientes tres palabras: POSITIVO, NEGATIVO o NEUTRO.
        """
        try:
            time.sleep(0.05)
            response = openai.ChatCompletion.create(
                model=OPENAI_MODEL_CLASIFICACION,
                messages=[
                    {"role": "system", "content": f"Tu única función es devolver POSITIVO, NEGATIVO o NEUTRO sobre la marca {MARCA_FIJA}."},
                    {"role": "user", "content": prompt}
                ],
                max_tokens=5,
                temperature=0.0
            )
            tono_raw = response.choices[0].message['content'].strip().upper()
            if 'POSITIVO' in tono_raw: final = 'Positivo'
            elif 'NEGATIVO' in tono_raw: final = 'Negativo'
            else: final = 'Neutro'
            self.cache_tonos[limpio] = final
            return final
        except Exception as e:
            st.warning(f"API Tono Warn: {str(e)[:100]}... Asignando Neutro.")
            return "Neutro (Error API)"

    def _agrupar_por_similitud(self, embeddings):
        n = len(embeddings)
        valid_indices = [i for i, emb in enumerate(embeddings) if emb is not None]
        if not valid_indices: return [[i] for i in range(n)]
        
        valid_embeddings = np.array([embeddings[i] for i in valid_indices])
        
        if len(valid_embeddings) > 1:
            clustering = AgglomerativeClustering(n_clusters=None, distance_threshold=1-self.high_similarity_threshold, metric='cosine', linkage='average').fit(valid_embeddings)
            labels = clustering.labels_
        else:
            labels = [0]

        grupos_dict = defaultdict(list)
        for i, label in enumerate(labels):
            original_index = valid_indices[i]
            grupos_dict[label].append(original_index)

        grupos_finales = list(grupos_dict.values())
        indices_con_grupo = set(valid_indices)
        for i in range(n):
            if i not in indices_con_grupo:
                grupos_finales.append([i])
        
        return grupos_finales

    def procesar_lote(self, df_columna_resumen, progress_bar):
        resumenes = df_columna_resumen.tolist()
        n = len(resumenes)
        tonos_finales = [""] * n
        
        embeddings = [_get_embedding_cached(r, self.cache_embeddings) for r in resumenes]
        grupos = self._agrupar_por_similitud(embeddings)
        
        processed_count = 0
        for grupo_indices in grupos:
            if not grupo_indices: continue
            representante_idx = grupo_indices[0]
            tono_grupo = self._clasificar_gpt(resumenes[representante_idx])
            for idx in grupo_indices:
                tonos_finales[idx] = tono_grupo
            processed_count += len(grupo_indices)
            progress_bar.progress(processed_count / n, text=f"Clasificando Tono: {processed_count}/{n}")
            
        return tonos_finales

class ClasificadorCategoriasPolicia:
    def __init__(self):
        self.cache_categorias = {}

    def _clasificar_gpt(self, texto_resumen):
        limpio = _limpiar_texto(texto_resumen)
        if not limpio: return "Indeterminado"
        if limpio in self.cache_categorias: return self.cache_categorias[limpio]
        
        prompt = f"""
        Tu tarea es clasificar la siguiente noticia en UNA de las 6 categorías definidas para la Policía Nacional. Analiza el núcleo de la noticia y elige la categoría más apropiada.

        CATEGORÍAS Y DEFINICIONES:
        1.  **Entorno**: Menciones referenciales a la institución, donde no es el actor principal.
        2.  **Gestión**: Noticias sobre actividades proactivas, campañas, trata de personas, jornadas de acompañamiento, consejos de seguridad, deportaciones, líneas de emergencia.
        3.  **Operativos**: Noticias cuyo foco principal son operativos policiales (capturas, incautaciones, etc.).
        4.  **Institucional**: Noticias sobre la estructura interna: nombramientos, presupuesto, investigaciones internas, irregularidades administrativas, corrupción.
        5.  **Delitos**: Noticias sobre hechos delictivos donde la policía no está ejecutando un operativo en el momento (ej. se reporta un robo, un homicidio, etc.).
        6.  **Centros de reclusión**: Noticias relacionadas específicamente con CAIs, estaciones de policía u otros centros de detención transitoria.

        NOTICIA: --- {limpio[:TEXT_TRUNCATION_LIMIT]} ---

        Responde ÚNICAMENTE con el nombre de la categoría (ej: Gestión).
        """
        try:
            time.sleep(0.05)
            response = openai.ChatCompletion.create(
                model=OPENAI_MODEL_CLASIFICACION,
                messages=[
                    {"role": "system", "content": "Eres un experto en clasificar noticias según categorías predefinidas."},
                    {"role": "user", "content": prompt}
                ],
                max_tokens=10,
                temperature=0.0
            )
            categoria = response.choices[0].message.content.strip()
            categorias_validas = ['Entorno', 'Gestión', 'Operativos', 'Institucional', 'Delitos', 'Centros de reclusión']
            if categoria not in categorias_validas:
                categoria = "Entorno" 
            self.cache_categorias[limpio] = categoria
            return categoria
        except Exception as e:
            st.warning(f"API Categoría Warn: {str(e)[:100]}... Asignando Entorno.")
            return "Entorno (Error API)"

    def procesar_lote(self, df_columna_resumen, progress_bar):
        resumenes = df_columna_resumen.tolist()
        n = len(resumenes)
        categorias_finales = []
        for i, resumen in enumerate(resumenes):
            categoria = self._clasificar_gpt(resumen)
            categorias_finales.append(categoria)
            progress_bar.progress((i + 1) / n, text=f"Clasificando Tema Principal: {i+1}/{n}")
        return categorias_finales

class ClasificadorSubtemas:
    def __init__(self):
        self.cache_embeddings = {}
        self.cache_subtemas_cluster = {}
        self.distance_threshold_resumenes = 0.35
        self.min_cluster_size = 2
        self.distance_threshold_consolidacion = 0.28

    def _generar_subtema_gpt(self, contexto, es_para_cluster=False, num_muestras=0):
        limpio = _limpiar_texto(contexto)
        if es_para_cluster and limpio in self.cache_subtemas_cluster:
            return self.cache_subtemas_cluster[limpio]

        instruccion = f"Este contexto representa un grupo de {num_muestras} noticias similares. Genera un subtema que refleje su núcleo semántico común." if es_para_cluster else "Genera un subtema que capture la esencia de este texto individual."
        prompt = f"""
        Tu tarea es analizar el siguiente contexto y generar un SUBTEMA PRINCIPAL que cumpla con estos criterios estrictos:
        - **Máximo 4 palabras.**
        - Frase nominal coherente y específica.
        - NO terminar en preposiciones o artículos (ej: de, la, y, en).

        {instruccion}

        Contexto: --- {limpio[:TEXT_TRUNCATION_LIMIT]} ---

        Responde ÚNICAMENTE con el subtema. Ejemplo válido: "Captura líder banda criminal".
        """
        try:
            time.sleep(0.05)
            response = openai.ChatCompletion.create(
                model=OPENAI_MODEL_CLASIFICACION,
                messages=[
                    {"role": "system", "content": "Eres un experto en síntesis temática precisa y concisa."},
                    {"role": "user", "content": prompt}
                ],
                max_tokens=15,
                temperature=0.1
            )
            subtema_bruto = response.choices[0].message.content.strip().replace('"', '')
            palabras_malas = ['de', 'la', 'los', 'las', 'un', 'una', 'y', 'e', 'o', 'para', 'con', 'en']
            palabras = subtema_bruto.split()
            if palabras and palabras[-1].lower() in palabras_malas:
                subtema_bruto = " ".join(palabras[:-1])
            
            final = subtema_bruto if subtema_bruto else "Subtema Genérico"
            if es_para_cluster:
                self.cache_subtemas_cluster[limpio] = final
            return final
        except Exception:
            return "Subtema (Error API)"

    def procesar_lote(self, df_columna_resumen, progress_bar):
        resumenes = [_limpiar_texto(r) for r in df_columna_resumen]
        n = len(resumenes)
        df_lote = pd.DataFrame({'resumen': resumenes, 'subtema': "Pendiente", 'cluster_id': -1})

        embeddings = [_get_embedding_cached(r, self.cache_embeddings) for r in resumenes]
        validos = [(i, emb) for i, emb in enumerate(embeddings) if emb is not None]

        if len(validos) < 2:
            for i, r in enumerate(resumenes): 
                df_lote.at[i, 'subtema'] = self._generar_subtema_gpt(r)
                progress_bar.progress((i + 1) / n, text=f"Clasificando Subtema (Individual): {i+1}/{n}")
            return df_lote['subtema'].tolist()

        indices, emb_matrix = zip(*validos)
        clustering = AgglomerativeClustering(n_clusters=None, distance_threshold=self.distance_threshold_resumenes, metric='cosine', linkage='complete').fit(np.array(emb_matrix))
        df_lote.loc[list(indices), 'cluster_id'] = clustering.labels_

        total_procesado = 0
        for cluster_id in df_lote[df_lote['cluster_id'] != -1]['cluster_id'].unique():
            cluster_df = df_lote[df_lote['cluster_id'] == cluster_id]
            if len(cluster_df) >= self.min_cluster_size:
                contexto_resumenes = sorted(cluster_df['resumen'].tolist(), key=len, reverse=True)[:5]
                contexto = " | ".join(contexto_resumenes)
                subtema_cluster = self._generar_subtema_gpt(contexto, es_para_cluster=True, num_muestras=len(cluster_df))
                df_lote.loc[cluster_df.index, 'subtema'] = subtema_cluster
            
            total_procesado += len(cluster_df)
            progress_bar.progress(total_procesado / (n * 2), text=f"Subtemas de Grupo: {total_procesado}/{n}")
        
        for i, row in df_lote.iterrows():
            if "Pendiente" in str(row['subtema']):
                df_lote.at[i, 'subtema'] = self._generar_subtema_gpt(row['resumen'])
            progress_bar.progress(0.5 + (i + 1) / (n * 2), text=f"Subtemas Individuales: {i+1}/{n}")

        subtemas_unicos = df_lote[~df_lote['subtema'].str.contains("Error|Pendiente", na=True)]['subtema'].unique().tolist()
        if len(subtemas_unicos) > 1:
            subtema_embeddings = [_get_embedding_cached(t, self.cache_embeddings) for t in subtemas_unicos]
            validos_temas = [(i, emb) for i, emb in enumerate(subtema_embeddings) if emb is not None]
            if len(validos_temas) > 1:
                indices_t, emb_matrix_t = zip(*validos_temas)
                clustering_t = AgglomerativeClustering(n_clusters=None, distance_threshold=self.distance_threshold_consolidacion, metric='cosine', linkage='average').fit(np.array(emb_matrix_t))
                mapa_consolidacion = {}
                temas_unicos_validos = [subtemas_unicos[i] for i in indices_t]
                
                for i, cluster_label in enumerate(clustering_t.labels_):
                    tema_original = temas_unicos_validos[i]
                    if tema_original not in mapa_consolidacion:
                        temas_en_cluster = [temas_unicos_validos[j] for j, lab in enumerate(clustering_t.labels_) if lab == cluster_label]
                        tema_canonico = sorted(temas_en_cluster, key=len)[0]
                        for t in temas_en_cluster:
                            mapa_consolidacion[t] = tema_canonico

                df_lote['subtema'] = df_lote['subtema'].map(mapa_consolidacion).fillna(df_lote['subtema'])
                
        return df_lote['subtema'].tolist()

# ==============================================================================
# FUNCIONES DE AYUDA Y PROCESAMIENTO
# ==============================================================================
def norm_key(text): return re.sub(r'\W+', '', str(text).lower().strip()) if text else ""

def extract_link(cell):
    if hasattr(cell, 'hyperlink') and cell.hyperlink: return {"value": "Link", "url": cell.hyperlink.target}
    if isinstance(cell.value, str):
        match = re.search(r'=HYPERLINK\("([^"]+)"', cell.value); 
        if match: return {"value": "Link", "url": match.group(1)}
    return {"value": cell.value, "url": None}

def normalize_title_for_comparison(title):
    if not isinstance(title, str): return ""
    cleaned_title = re.sub(r'\s*\|\s*.+$', '', title).strip()
    return re.sub(r'\W+', ' ', cleaned_title).lower().strip()

def clean_title_for_output(title): return re.sub(r'\s*\|\s*[\w\s]+$', '', str(title)).strip()

def corregir_texto(text):
    if not isinstance(text, str): return text
    text = re.sub(r'(<br>|\[\.\.\.\]|\s+)', ' ', text).strip()
    match = re.search(r'[A-Z]', text); 
    if match: text = text[match.start():]
    if text and not text.endswith('...'): text = text.rstrip('.') + '...'
    return text

def are_duplicates(row1, row2, key_map, title_similarity_threshold=0.85, date_proximity_days=1):
    if norm_key(row1.get(key_map['menciones'])) != norm_key(row2.get(key_map['menciones'])): return False
    if norm_key(row1.get(key_map['medio'])) != norm_key(row2.get(key_map['medio'])): return False

    tipo_medio1 = norm_key(row1.get(key_map['tipo_medio']))
    titulo1 = normalize_title_for_comparison(row1.get(key_map['titulo']))
    titulo2 = normalize_title_for_comparison(row2.get(key_map['titulo']))

    try:
        fecha1 = pd.to_datetime(row1.get(key_map['fecha'])).date()
        fecha2 = pd.to_datetime(row2.get(key_map['fecha'])).date()
    except (ValueError, TypeError): return False

    if tipo_medio1 == 'internet':
        if row1.get(key_map['hora']) == row2.get(key_map['hora']): return False
        if abs((fecha1 - fecha2).days) > date_proximity_days: return False
        if titulo1 and titulo1 == titulo2: return True
        if SequenceMatcher(None, titulo1, titulo2).ratio() >= title_similarity_threshold: return True
    else:
        if norm_key(row1.get(key_map['tipo_medio'])) in {'radio', 'televisión'}:
             if row1.get(key_map['hora']) != row2.get(key_map['hora']): return False
        if fecha1 != fecha2: return False
        if titulo1 and titulo1 == titulo2: return True
            
    return False
# ==============================================================================
# LÓGICA PRINCIPAL DE PROCESAMIENTO
# ==============================================================================

def run_dossier_logic(sheet):
    headers = [cell.value for cell in sheet[1] if cell.value]
    norm_keys = [norm_key(h) for h in headers]
    key_map = {nk: nk for nk in norm_keys}
    key_map.update({
        'titulo': norm_key('Título'), 'resumen': norm_key('Resumen - Aclaracion'), 
        'tipo_medio': norm_key('Tipo de Medio'), 'medio': norm_key('Medio'), 
        'fecha': norm_key('Fecha'), 'hora': norm_key('Hora'), 
        'menciones': norm_key('Menciones - Empresa'), 'link_nota': norm_key('Link Nota'), 
        'link_streaming': norm_key('Link (Streaming - Imagen)'), 'tono': norm_key('Tono'), 
        'tema': norm_key('Tema'), 'subtema': norm_key('Subtema'), 'region': norm_key('Región'), 
        'tonoai': norm_key('Tono AI'), 'dimension': norm_key('Dimensión'), 
        'duracion': norm_key('Duración - Nro. Caracteres')
    })
    
    processed_rows = []
    for row in sheet.iter_rows(min_row=2):
        if all(c.value is None for c in row): continue
        base_data = {norm_keys[i]: extract_link(cell) if norm_keys[i] in [key_map['link_nota'], key_map['link_streaming']] else cell.value for i, cell in enumerate(row) if i < len(norm_keys)}
        
        tm_norm = norm_key(base_data.get(key_map['tipo_medio'])); tipo_medio_val = base_data.get(key_map['tipo_medio'])
        if tm_norm in {'aire', 'cable'}: tipo_medio_val = 'Televisión'
        elif tm_norm in {'am', 'fm'}: tipo_medio_val = 'Radio'
        elif tm_norm == 'diario': tipo_medio_val = 'Prensa'
        elif tm_norm == 'online': tipo_medio_val = 'Internet'
        elif tm_norm == 'revista': tipo_medio_val = 'Revista'
        base_data[key_map['tipo_medio']] = tipo_medio_val
        
        link_nota = base_data.get(key_map['link_nota']); link_streaming = base_data.get(key_map['link_streaming'])
        if tipo_medio_val == "Internet": base_data[key_map['link_nota']], base_data[key_map['link_streaming']] = link_streaming, link_nota
        elif tipo_medio_val in {"Prensa", "Revista"}:
            if (not link_nota or not link_nota.get('url')) and (link_streaming and link_streaming.get('url')): base_data[key_map['link_nota']] = link_streaming
            base_data[key_map['link_streaming']] = None
        elif tipo_medio_val in {"Radio", "Televisión"}: base_data[key_map['link_streaming']] = None

        if tipo_medio_val in {"Radio", "Televisión"} and key_map.get('duracion') in base_data and key_map.get('dimension') in base_data:
            base_data[key_map['dimension']] = base_data.get(key_map['duracion']); base_data[key_map['duracion']] = None
                
        menciones = [m.strip() for m in str(base_data.get(key_map['menciones']) or '').split(';') if m.strip()]
        if not menciones: processed_rows.append(base_data)
        else:
            for mencion in menciones: new_row = deepcopy(base_data); new_row[key_map['menciones']] = mencion; processed_rows.append(new_row)
            
    for idx, row in enumerate(processed_rows): row.update({'original_index': idx, 'is_duplicate': False})
    
    processed_rows.sort(key=lambda r: ('"' not in str(r.get(key_map['titulo'], '')), pd.to_datetime(r.get(key_map['fecha']), errors='coerce'), r['original_index']))
    
    for i in range(len(processed_rows)):
        if processed_rows[i]['is_duplicate']: continue
        for j in range(i + 1, len(processed_rows)):
            row1, row2 = processed_rows[i], processed_rows[j]
            if row2['is_duplicate']: continue 
            if are_duplicates(row1, row2, key_map):
                processed_rows[j]['is_duplicate'] = True

    processed_rows.sort(key=lambda r: r['original_index'])

    for row in processed_rows:
        if row['is_duplicate']:
            row[key_map['tono']] = "Duplicada"; row[key_map['tema']] = "-"; row[key_map['subtema']] = "-"; row[key_map['tonoai']] = "-"
            
    return processed_rows, key_map

def generate_output_excel(all_processed_rows, key_map):
    out_wb = Workbook()
    out_sheet = out_wb.active
    out_sheet.title = "Resultado"

    final_order = [
        "ID Noticia", "Fecha", "Hora", "Medio", "Tipo de Medio", "Sección - Programa", 
        "Región", "Título", "Autor - Conductor", "Nro. Pagina", "Dimensión", 
        "Duración - Nro. Caracteres", "CPE", "Tier", "Audiencia", 
        "Tono", "Tono AI", "Tema", "Subtema",
        "Resumen - Aclaracion", "Link Nota", 
        "Link (Streaming - Imagen)", "Menciones - Empresa"
    ]
    out_sheet.append(final_order)

    link_style = NamedStyle(name="Hyperlink_Custom", font=Font(color="0000FF", underline="single"))
    if "Hyperlink_Custom" not in out_wb.named_styles: out_wb.add_named_style(link_style)

    link_nota_idx = final_order.index("Link Nota") + 1 if "Link Nota" in final_order else -1
    link_streaming_idx = final_order.index("Link (Streaming - Imagen)") + 1 if "Link (Streaming - Imagen)" in final_order else -1

    for row_data in all_processed_rows:
        if not row_data.get('is_duplicate'):
            row_data[key_map['titulo']] = clean_title_for_output(row_data.get(key_map['titulo']))
        row_data[key_map['resumen']] = corregir_texto(row_data.get(key_map['resumen']))
        
        row_to_append = []
        for header in final_order:
            key = norm_key(header)
            val = row_data.get(key)
            row_to_append.append(val['value'] if isinstance(val, dict) and 'value' in val else val)
        out_sheet.append(row_to_append)
        
        current_row_idx = out_sheet.max_row
        
        if link_nota_idx != -1 and isinstance(row_data.get(key_map['link_nota']), dict) and row_data.get(key_map['link_nota'], {}).get("url"):
            cell = out_sheet.cell(row=current_row_idx, column=link_nota_idx)
            cell.hyperlink = row_data[key_map['link_nota']]["url"]; cell.value = "Link"; cell.style = "Hyperlink_Custom"
        
        if link_streaming_idx != -1 and isinstance(row_data.get(key_map['link_streaming']), dict) and row_data.get(key_map['link_streaming'], {}).get("url"):
            cell = out_sheet.cell(row=current_row_idx, column=link_streaming_idx)
            cell.hyperlink = row_data[key_map['link_streaming']]["url"]; cell.value = "Link"; cell.style = "Hyperlink_Custom"

    output = io.BytesIO(); out_wb.save(output); output.seek(0)
    return output.getvalue()

def run_full_process(dossier_file, region_file, internet_file):
    try: openai.api_key = st.secrets["OPENAI_API_KEY"]
    except Exception: st.error("Error: 'OPENAI_API_KEY' no encontrado en los Secrets de Streamlit."); st.stop()

    with st.status("Paso 1/6: Limpiando y deduplicando...", expanded=True) as status:
        wb = load_workbook(dossier_file, data_only=False)
        all_processed_rows, key_map = run_dossier_logic(wb.active)
        num_duplicates = sum(1 for r in all_processed_rows if r.get('is_duplicate'))
        to_analyze_count = len(all_processed_rows) - num_duplicates
        status.update(label=f"✅ Limpieza completada. {to_analyze_count} noticias únicas de {len(all_processed_rows)} filas totales.", state="complete")

    with st.status("Paso 2/6: Aplicando mapeos (Región e Internet)...", expanded=True) as status:
        df_region = pd.read_excel(region_file); region_map = pd.Series(df_region.iloc[:, 1].values, index=df_region.iloc[:, 0].astype(str).str.lower().str.strip()).to_dict()
        df_internet = pd.read_excel(internet_file); internet_map = pd.Series(df_internet.iloc[:, 1].values, index=df_internet.iloc[:, 0].astype(str).str.lower().str.strip()).to_dict()
        
        for row in all_processed_rows:
            medio_key = str(row.get(key_map['medio'], '')).lower().strip()
            row[key_map['region']] = region_map.get(medio_key, "Error")
            if row.get(key_map['tipo_medio']) == 'Internet':
                row[key_map['medio']] = internet_map.get(medio_key, row.get(key_map['medio']))
        
        status.update(label="✅ Mapeos aplicados.", state="complete")

    rows_to_analyze = [row for row in all_processed_rows if not row.get('is_duplicate')]
    if rows_to_analyze:
        for row in rows_to_analyze: row['resumen_api'] = str(row.get(key_map['titulo'], '')) + ". " + str(row.get(key_map['resumen'], ''))
        df_temp_api = pd.DataFrame(rows_to_analyze)
        
        with st.status(f"Paso 3/6: Clasificando Tono para {len(rows_to_analyze)} noticias...", expanded=True) as status:
            clasificador_tono = ClasificadorTonoPolicia()
            progress_bar_tono = st.progress(0, "Iniciando análisis de Tono...")
            df_temp_api[key_map['tonoai']] = clasificador_tono.procesar_lote(df_temp_api['resumen_api'], progress_bar_tono)
            status.update(label="✅ Clasificación de Tono completada.", state="complete")

        with st.status(f"Paso 4/6: Clasificando Tema Principal para {len(rows_to_analyze)} noticias...", expanded=True) as status:
            clasificador_categoria = ClasificadorCategoriasPolicia()
            progress_bar_cat = st.progress(0, "Iniciando análisis de Tema...")
            df_temp_api[key_map['tema']] = clasificador_categoria.procesar_lote(df_temp_api['resumen_api'], progress_bar_cat)
            status.update(label="✅ Clasificación de Tema Principal completada.", state="complete")
        
        with st.status(f"Paso 5/6: Generando Subtemas para {len(rows_to_analyze)} noticias...", expanded=True) as status:
            clasificador_subtema = ClasificadorSubtemas()
            progress_bar_subtema = st.progress(0, "Iniciando análisis de Subtema...")
            df_temp_api[key_map['subtema']] = clasificador_subtema.procesar_lote(df_temp_api['resumen_api'], progress_bar_subtema)
            status.update(label="✅ Generación de Subtemas completada.", state="complete")

        results_map = df_temp_api.set_index('original_index')[[key_map['tonoai'], key_map['tema'], key_map['subtema']]].to_dict('index')
        for row in all_processed_rows:
            if not row.get('is_duplicate'):
                result = results_map.get(row['original_index'])
                if result: 
                    row[key_map['tonoai']] = result[key_map['tonoai']]
                    row[key_map['tema']] = result[key_map['tema']]
                    row[key_map['subtema']] = result[key_map['subtema']]

    with st.status("Paso 6/6: Generando archivo de salida...", expanded=True) as status:
        st.session_state['output_data'] = generate_output_excel(all_processed_rows, key_map)
        st.session_state['output_filename'] = f"Informe_IA_Policia_Nacional_{datetime.datetime.now().strftime('%Y%m%d')}.xlsx"
        st.session_state['processing_complete'] = True
        status.update(label="✅ Archivo de salida listo.", state="complete")

# ==============================================================================
# INICIO Y ESTRUCTURA DE LA APP
# ==============================================================================
if check_password():
    st.title(f"🤖 Análisis de Noticias con IA para {MARCA_FIJA}")
    st.markdown("Esta herramienta limpia, deduplica y analiza noticias (Tono, Tema y Subtema) usando la API de OpenAI.")

    if not st.session_state.get('processing_complete', False):
        with st.form("input_form"):
            st.header("Carga de Archivos")
            st.info(f"El análisis se realizará específicamente para la marca: **{MARCA_FIJA}**.")
            dossier_file = st.file_uploader("1. Archivo de Dossier (Principal)", type=["xlsx"])
            region_file = st.file_uploader("2. Archivo de Mapeo de Región", type=["xlsx"])
            internet_file = st.file_uploader("3. Archivo de Mapeo de Internet", type=["xlsx"])
            submitted = st.form_submit_button("🚀 Iniciar Análisis Completo")
            
            if submitted:
                if not all([dossier_file, region_file, internet_file]):
                    st.warning("Por favor, carga los tres archivos requeridos para continuar.")
                else:
                    st.session_state.clear()
                    st.session_state.password_correct = True
                    run_full_process(dossier_file, region_file, internet_file)
                    st.rerun()
    else:
        st.success("🎉 ¡Proceso finalizado con éxito!")
        st.info("El informe con los datos procesados y el análisis de Tono, Tema y Subtema está listo para ser descargado.")
        st.download_button(
            label="📥 Descargar Informe Completo",
            data=st.session_state['output_data'],
            file_name=st.session_state['output_filename']
        )
        if st.button("🧹 Empezar Nuevo Análisis"):
            st.session_state.clear()
            st.session_state.password_correct = True
            st.rerun()
